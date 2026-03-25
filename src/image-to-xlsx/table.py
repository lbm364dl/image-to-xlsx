import re

import numpy as np
from collections import defaultdict
from definitions import (
    MISSPELLINGS,
    MISSPELLINGS_REGEX,
    NOT_NUMBER,
    ONE_NUMBER,
    AT_LEAST_TWO_NUMBERS,
)
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.comments import Comment
from openpyxl.styles import Border, PatternFill, Side
from utils import get_cell_color, split_footnotes


class Table:
    def __init__(self, whole_image, page, table_bbox):
        self.page = page
        self.footer_text = None
        self.image = whole_image
        self.table_bbox = table_bbox

    # ------------------------------------------------------------------
    # Populating table_data from service responses
    # ------------------------------------------------------------------

    def set_table_from_service_response(self, cells):
        """Populate table_data from the unified extraction service format.

        ``cells`` is a list of dicts, each with keys:
        ``row``, ``col``, ``text``, ``confidence``.
        """
        self.table_data = defaultdict(lambda: defaultdict(list))
        for cell in cells:
            self.table_data[cell["row"]][cell["col"]].append(
                {"text": cell.get("text", ""), "confidence": cell.get("confidence")}
            )

    # ------------------------------------------------------------------
    # Numeric helpers
    # ------------------------------------------------------------------

    def is_numeric_cell(self, text, threshold=0.6):
        numeric = sum("0" <= c <= "9" for c in text)
        return numeric >= threshold * len(text)

    def cell_to_float(self, value):
        precision = self.page.document.fixed_decimal_places
        return float(value) / 10**precision

    def maybe_parse_numeric_cell(
        self, cell, decimal_separator, thousands_separator, fix_num_misspellings
    ):
        text = cell["text"]
        if not self.is_numeric_cell(text):
            return text, NOT_NUMBER

        text = text.replace(thousands_separator, "")
        if decimal_separator == ",":
            one_number = cell["cnt_commas"] <= 1
            text = text.replace(decimal_separator, ".")
        else:
            one_number = cell["cnt_dots"] <= 1

        if fix_num_misspellings:
            text = MISSPELLINGS_REGEX.sub(
                lambda x: str(MISSPELLINGS[x.group()]), text.upper()
            )

        try:
            return self.cell_to_float(text), NOT_NUMBER
        except ValueError:
            if not one_number:
                return text, AT_LEAST_TWO_NUMBERS if text else NOT_NUMBER
            maybe_sign = "-" if text.startswith("-") else ""
            only_numeric = maybe_sign + re.sub(r"[^.0-9]", "", text)

            if only_numeric:
                try:
                    return self.cell_to_float(only_numeric), ONE_NUMBER
                except ValueError:
                    return only_numeric, NOT_NUMBER
            else:
                return "", NOT_NUMBER

    # ------------------------------------------------------------------
    # Matrix building / postprocessing
    # ------------------------------------------------------------------

    def extend_rows(self):
        split_table = []
        for row in self.table_data.values():
            rows = []
            for j, col in row.items():
                for i, part in enumerate(col):
                    if len(rows) <= i:
                        rows.append(defaultdict(list))
                    rows[i][j].append(part)
                    assert len(rows[i][j]) == 1
            split_table += rows

        self.table_data = defaultdict(
            lambda: defaultdict(list),
            {i: row for i, row in enumerate(split_table)},
        )

    def join_cell_parts(self, cell):
        texts = [cell_part["text"] for cell_part in cell]
        conf_arr = np.array(
            [
                cell_part["confidence"]
                for cell_part in cell
                if cell_part["confidence"] is not None
            ]
        )
        confidence = conf_arr.mean() if conf_arr.size > 0 else None
        return {
            "text": " ".join(texts),
            "confidence": (
                confidence
                if confidence is not None and not np.isnan(confidence)
                else None
            ),
        }

    def clean_cell_text(self, cell_text, remove_dots_and_commas):
        cell_text = ILLEGAL_CHARACTERS_RE.sub(r"", cell_text)
        dots = sum(c == "." for c in cell_text)
        commas = sum(c == "," for c in cell_text)

        if remove_dots_and_commas:
            cell_text = cell_text.replace(".", "").replace(",", "")

        return cell_text, dots, commas

    def as_clean_matrix(self, remove_dots_and_commas):
        n = max([i + 1 for i in self.table_data.keys()], default=0)
        m = max(
            [j + 1 for cols in self.table_data.values() for j in cols.keys()],
            default=0,
        )
        table_data = [
            [
                {
                    "text": "",
                    "confidence": None,
                    "footnotes": [],
                    "cnt_numbers": 0,
                    "cnt_dots": 0,
                    "cnt_commas": 0,
                }
                for _ in range(m)
            ]
            for _ in range(n)
        ]

        for row, cols in self.table_data.items():
            for col, cell in cols.items():
                cell = self.join_cell_parts(cell)
                cell["text"], cell["cnt_dots"], cell["cnt_commas"] = (
                    self.clean_cell_text(cell["text"], remove_dots_and_commas)
                )
                cell["text"], cell["footnotes"] = split_footnotes(cell["text"])
                table_data[row][col] = {**cell, "cnt_numbers": 0}

        return table_data

    def overwrite_seminumeric_cells_confidence(
        self,
        table_matrix,
        decimal_separator,
        thousands_separator,
        fix_num_misspellings,
    ):
        for i, row in enumerate(table_matrix):
            for j, col in enumerate(row):
                _, cnt_numbers = self.maybe_parse_numeric_cell(
                    col,
                    decimal_separator,
                    thousands_separator,
                    fix_num_misspellings,
                )
                table_matrix[i][j]["cnt_numbers"] = cnt_numbers
        return table_matrix

    def maybe_parse_numeric_cells(
        self,
        table_matrix,
        decimal_separator,
        thousands_separator,
        fix_num_misspellings,
    ):
        for i, row in enumerate(table_matrix):
            for j, col in enumerate(row):
                table_matrix[i][j]["text"], _ = self.maybe_parse_numeric_cell(
                    col,
                    decimal_separator,
                    thousands_separator,
                    fix_num_misspellings,
                )
        return table_matrix

    # ------------------------------------------------------------------
    # Excel output
    # ------------------------------------------------------------------

    def add_to_sheet(self, page_num, table_num, table_matrix, footer_text):
        thin_white_border = Border(
            left=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
            top=Side(style="thin", color="FFFFFF"),
            bottom=Side(style="thin", color="FFFFFF"),
        )
        thick_blue_border = Border(
            left=Side(style="thick", color="0000FF"),
            right=Side(style="thick", color="0000FF"),
            top=Side(style="thick", color="0000FF"),
            bottom=Side(style="thick", color="0000FF"),
        )
        thick_violet_border = Border(
            left=Side(style="thick", color="FF00FF"),
            right=Side(style="thick", color="FF00FF"),
            top=Side(style="thick", color="FF00FF"),
            bottom=Side(style="thick", color="FF00FF"),
        )

        self.page.document.footers_workbook.active.append(
            [page_num, table_num, footer_text]
        )

        sheet_name = f"page_{page_num}_table_{table_num}"
        sheet = self.page.document.workbook.create_sheet(sheet_name)

        for i, row in enumerate(table_matrix, 1):
            for j, col in enumerate(row, 1):
                cell = sheet.cell(row=i, column=j, value=col["text"])
                cell_color = get_cell_color(col["confidence"])
                if cell_color:
                    cell.fill = PatternFill(
                        start_color=cell_color, fill_type="solid"
                    )
                    cell.border = [
                        thin_white_border,
                        thick_violet_border,
                        thick_blue_border,
                    ][col["cnt_numbers"]]

                    if col["footnotes"]:
                        cell.comment = Comment(
                            ",".join(col["footnotes"]), "automatic"
                        )

    # ------------------------------------------------------------------
    # NLP postprocessing (OpenAI)
    # ------------------------------------------------------------------

    def nlp_postprocess(
        self, table_matrix, text_language="en", nlp_postprocess_prompt_file=None
    ):
        from postprocessing import nlp_clean

        return nlp_clean(table_matrix, text_language, nlp_postprocess_prompt_file)
